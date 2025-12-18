import os
import pandas as pd
import docx
import PyPDF2
import numpy as np
from typing import List, Dict, Optional, Tuple, Any

class RequirementParser:
    """Parser for extracting requirements from different document formats"""
    
    def __init__(self):
        self.supported_extensions = ['.docx', '.pdf', '.xlsx', '.csv']
    
    def is_supported_file(self, file_path: str) -> bool:
        """Check if the file format is supported"""
        _, extension = os.path.splitext(file_path)
        return extension.lower() in self.supported_extensions
    
    def extract_requirements(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract requirements from the document based on file type"""
        _, extension = os.path.splitext(file_path)
        extension = extension.lower()
        print(f"inside extract_requirements: Extracting requirements from {file_path} with extension: {extension}")
        
        if extension == '.docx':
            return self._extract_from_docx(file_path)
        elif extension == '.pdf':
            return self._extract_from_pdf(file_path)
        elif extension == '.xlsx':
            return self._extract_from_excel(file_path)
        elif extension == '.csv':
            return self._extract_from_csv(file_path)
        else:
            raise ValueError(f"Unsupported file format: {extension}")
    
    def _extract_from_docx(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract requirements from Word documents (.docx)"""
        doc = docx.Document(file_path)
        requirements = []
        
        # Extract from tables
        for table in doc.tables:
            headers = []
            # Get headers from first row
            for cell in table.rows[0].cells:
                headers.append(cell.text.strip())
            
            # Check if this table contains requirements
            if 'id' in [h.lower() for h in headers] and any(h.lower() in ['requirement', 'text', 'description'] for h in headers):
                # Process each row as a requirement
                for i, row in enumerate(table.rows):
                    if i == 0:  # Skip header row
                        continue
                    
                    req = {}
                    for j, cell in enumerate(row.cells):
                        if j < len(headers):  # Ensure we don't go out of bounds
                            req[headers[j]] = cell.text.strip()
                    
                    # Only add if it has both ID and text
                    if req and 'ID' in req and any(key in req for key in ['Requirement', 'Text', 'Description']):
                        requirements.append(req)
        
        return requirements
    
    def _extract_from_pdf(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract requirements from PDF documents"""
        # Basic PDF parsing - this is a placeholder and would need more sophisticated
        # parsing for real-world PDFs with tables
        requirements = []
        
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            # This is a very simplified approach - in reality, PDF table extraction
            # would need more sophisticated libraries like tabula-py or camelot
            # For now, we'll just note that PDF parsing is implemented
            
        return requirements
    
    def _extract_from_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract requirements from Excel spreadsheets"""
        requirements = []
        
        try:
            # Normalize the file path
            file_path = os.path.abspath(file_path)
            print(f"Normalized file path: {file_path}")
            
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Excel file not found at: {file_path}")
            
            # First try to open the Excel file
            print(f"Attempting to open Excel file: {file_path}")
            
            # Read the Excel file using openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(filename=file_path, read_only=True)
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                print(f"\nProcessing sheet: {sheet_name}")
                ws = wb[sheet_name]
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).strip() if cell.value else '')
                
                print(f"Found headers: {headers}")  # Debug print
                
                # Find required columns (case-insensitive)
                id_col_idx = None
                text_col_idx = None
                
                for idx, header in enumerate(headers):
                    header_lower = header.lower()
                    if header_lower == 'id':
                        id_col_idx = idx
                    elif header_lower in ['requirement', 'text', 'description']:
                        text_col_idx = idx
                
                print(f"ID column index: {id_col_idx}, Text column index: {text_col_idx}")  # Debug print
                
                if id_col_idx is None:
                    print(f"Warning: 'ID' column not found in sheet {sheet_name}")
                    continue
                
                if text_col_idx is None:
                    print(f"Warning: Could not find requirement text column in sheet {sheet_name}")
                    continue
                
                # Process each row
                for row in ws.iter_rows(min_row=2):  # Skip header row
                    try:
                        # Get raw values
                        id_value = row[id_col_idx].value
                        text_value = row[text_col_idx].value
                        
                        # Convert to string, handling numpy arrays and None values
                        req_id = str(id_value) if id_value is not None else ""
                        req_text = str(text_value) if text_value is not None else ""
                        
                        print(f"Processing row - ID: {req_id}, Text: {req_text}")  # Debug print
                        
                        if req_id and req_text and not req_id.startswith(('nan', 'None')):
                            requirements.append({
                                'ID': req_id,
                                'Text': req_text
                            })
                    except Exception as e:
                        print(f"Error processing row: {str(e)}")
                        continue
            
            print(f"Extracted {len(requirements)} requirements")
            
        except Exception as e:
            print(f"Error processing Excel file: {str(e)}")
            print(f"Exception type: {type(e)}")
            print(f"Exception args: {e.args}")
            raise ValueError(f"Could not read Excel file: {str(e)}")
                
        return requirements
    
    def _extract_from_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract requirements from CSV files"""
        df = pd.read_csv(file_path)
        
        # Check if this CSV might contain requirements
        if 'ID' in df.columns and any(col in df.columns for col in ['Requirement', 'Text', 'Description']):
            # Convert to dict
            requirements = df.to_dict(orient='records')
            return requirements
        
        return [] 

